"""
Add MV{year} and MV{year-1} columns to each _with_market_values.xlsx file.

For the 2018 file: adds MV18 (avg of 2018 MV snapshots) and MV17 (avg of 2017 snapshots).
Values are in millions of euros (e.g., 8.0 = €8m, 0.8 = €800k).

Each cell contains an Excel formula with the individual values embedded, e.g. =(8+8)/2,
so the formula bar shows the breakdown while the cell displays the computed result.
Rows with no snapshots in that year get an empty cell (no formula).

Usage:
    py -3.11 Code/add_mv_columns.py                  # all years 2018-2025
    py -3.11 Code/add_mv_columns.py --years 2018 2023
"""

import argparse
import ast
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / 'Data'


def _parse_mv_str(s: str) -> float | None:
    """Parse value strings like '€8.00m' → 8.0, '€800k' → 0.8 (millions)."""
    cleaned = re.sub(r'[^\d.mk]', '', str(s).lower())
    try:
        if cleaned.endswith('m'):
            return float(cleaned[:-1])
        elif cleaned.endswith('k'):
            return float(cleaned[:-1]) / 1000
        elif cleaned:
            return float(cleaned)
    except ValueError:
        pass
    return None


def _extract_history(raw) -> list:
    """Return marketValueHistory list from a raw cell value (dict or string repr)."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return []
    if isinstance(raw, dict):
        return raw.get('marketValueHistory', [])
    try:
        s = re.sub(r'datetime\.datetime\([^)]+\)', 'None', str(raw))
        data = ast.literal_eval(s)
        if isinstance(data, dict):
            return data.get('marketValueHistory', [])
    except Exception:
        pass
    return []


def _year_values(history: list, year: int) -> list[float]:
    """Return all MV values (millions) from snapshots recorded in the given calendar year."""
    values = []
    for entry in history:
        date_str = str(entry.get('date', ''))
        try:
            # Dates are DD/MM/YYYY — year is the last segment
            entry_year = int(date_str.split('/')[-1]) if '/' in date_str else int(date_str[-4:])
        except (ValueError, IndexError):
            continue
        if entry_year != year:
            continue
        v = _parse_mv_str(entry.get('value', ''))
        if v is not None:
            values.append(v)
    return values


def _fmt_val(v: float) -> str:
    """Format a float cleanly: 8.0 → '8', 36.667 → '36.667'."""
    if v == int(v):
        return str(int(v))
    return str(round(v, 3))


def _make_formula(values: list[float]) -> str | None:
    """Build an Excel formula like '=(8+8)/2'. Returns None if no values."""
    if not values:
        return None
    if len(values) == 1:
        return f"={_fmt_val(values[0])}"
    parts = '+'.join(_fmt_val(v) for v in values)
    return f"=({parts})/{len(values)}"


def add_mv_columns(year: int) -> None:
    filepath = DATA_DIR / f'UEFA Stats {year}_with_market_values.xlsx'
    if not filepath.exists():
        print(f"  [skip] {filepath.name} not found")
        return

    print(f"Processing {filepath.name} ...")
    df = pd.read_excel(filepath)

    col_cur  = f'MV{str(year)[2:]}'       # e.g. 'MV18'
    col_prev = f'MV{str(year - 1)[2:]}'   # e.g. 'MV17'

    formulas_cur  = []
    formulas_prev = []

    for _, row in df.iterrows():
        history = _extract_history(row.get('Market Value History'))
        formulas_cur.append(_make_formula(_year_values(history, year)))
        formulas_prev.append(_make_formula(_year_values(history, year - 1)))

    # Add placeholder columns so they appear when pandas writes the sheet
    df[col_cur]  = None
    df[col_prev] = None
    df.to_excel(filepath, index=False)

    # Re-open with openpyxl to write formula strings into the new columns
    wb = load_workbook(filepath)
    ws = wb.active
    headers = {cell.value: cell.column for cell in ws[1]}
    col_cur_idx  = headers.get(col_cur)
    col_prev_idx = headers.get(col_prev)

    for row_idx, (f_cur, f_prev) in enumerate(zip(formulas_cur, formulas_prev), start=2):
        if f_cur and col_cur_idx:
            ws.cell(row=row_idx, column=col_cur_idx, value=f_cur)
        if f_prev and col_prev_idx:
            ws.cell(row=row_idx, column=col_prev_idx, value=f_prev)

    wb.save(filepath)

    filled_cur  = sum(1 for f in formulas_cur  if f)
    filled_prev = sum(1 for f in formulas_prev if f)
    print(f"  {col_cur}:  {filled_cur}/{len(df)} filled")
    print(f"  {col_prev}: {filled_prev}/{len(df)} filled")


def main():
    parser = argparse.ArgumentParser(description="Add MV{year} and MV{year-1} columns to market value files.")
    parser.add_argument('--years', nargs='+', type=int, default=list(range(2018, 2026)),
                        help='Years to process (default: 2018-2025)')
    args = parser.parse_args()

    print("Adding MV year columns to _with_market_values.xlsx files\n")
    for year in args.years:
        add_mv_columns(year)
    print("\nDone.")


if __name__ == '__main__':
    main()
